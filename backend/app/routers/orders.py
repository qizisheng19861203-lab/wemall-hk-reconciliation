from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import and_, or_, func
from typing import List, Optional
from datetime import datetime, timedelta
from app.database import get_db
from app.models.order import Order, OrderItem, ShippingStatus
from app.models.user import User
from app.core.deps import get_current_user, require_admin_or_operator, require_admin
from app.schemas.order import OrderResponse, OrderUpdate, OrderFilter
from app.services.wemall_api import WemallAPI

router = APIRouter(prefix="/orders", tags=["订单管理"])


def _active_store_filter(db: Session):
    from app.models.wemall_store_config import WemallStoreConfig
    store = db.query(WemallStoreConfig).filter(WemallStoreConfig.is_active == True).first()
    if store:
        return Order.wemall_store_id == store.id
    return True


def _build_query(db: Session, f: OrderFilter):
    q = db.query(Order).options(joinedload(Order.items).joinedload(OrderItem.product)).filter(_active_store_filter(db))
    q = q.filter(Order.is_test == False)  # 测试订单不在订单管理列表显示
    if f.start_date:
        q = q.filter(Order.order_date >= f.start_date)
    if f.end_date:
        q = q.filter(Order.order_date <= f.end_date)
    if f.shipping_status:
        q = q.filter(Order.shipping_status == f.shipping_status)
    if f.is_refunded is not None:
        q = q.filter(Order.is_refunded == f.is_refunded)
    if f.settlement_id:
        q = q.filter(Order.settlement_id == f.settlement_id)
    if f.unsettled_only:
        q = q.filter(Order.settlement_id == None)
    if f.keyword:
        q = q.filter(
            or_(
                Order.wemall_order_id.contains(f.keyword),
                Order.buyer_name.contains(f.keyword),
                Order.buyer_phone.contains(f.keyword),
            )
        )
    return q


@router.get("", response_model=List[OrderResponse])
def list_orders(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    shipping_status: Optional[ShippingStatus] = None,
    is_refunded: Optional[bool] = None,
    settlement_id: Optional[int] = None,
    unsettled_only: bool = False,
    keyword: Optional[str] = None,
    skip: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    f = OrderFilter(
        start_date=start_date, end_date=end_date,
        shipping_status=shipping_status, is_refunded=is_refunded,
        settlement_id=settlement_id, unsettled_only=unsettled_only,
        keyword=keyword,
    )
    return _build_query(db, f).order_by(Order.order_date.desc()).offset(skip).limit(limit).all()


@router.get("/count")
def count_orders(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    shipping_status: Optional[ShippingStatus] = None,
    is_refunded: Optional[bool] = None,
    settlement_id: Optional[int] = None,
    unsettled_only: bool = False,
    keyword: Optional[str] = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """返回当前筛选条件下的订单总数（用于前端分页显示真实页数）"""
    f = OrderFilter(
        start_date=start_date, end_date=end_date,
        shipping_status=shipping_status, is_refunded=is_refunded,
        settlement_id=settlement_id, unsettled_only=unsettled_only,
        keyword=keyword,
    )
    total = _build_query(db, f).order_by(None).count()
    return {"total": total}


_SHIP_LABEL = {"pending": "待发货", "shipped": "已发货", "delivered": "已签收", "returned": "已退货"}


@router.get("/export")
def export_orders(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    shipping_status: Optional[ShippingStatus] = None,
    is_refunded: Optional[bool] = None,
    settlement_id: Optional[int] = None,
    unsettled_only: bool = False,
    supply_only: bool = False,
    keyword: Optional[str] = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """按当前筛选条件导出订单明细 Excel（不分页，导全部）。每个商品一行。"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    f = OrderFilter(
        start_date=start_date, end_date=end_date,
        shipping_status=shipping_status, is_refunded=is_refunded,
        settlement_id=settlement_id, unsettled_only=unsettled_only,
        keyword=keyword,
    )
    orders = _build_query(db, f).order_by(Order.order_date.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "订单明细"
    headers = [
        "订单号", "下单时间(北京)", "收件人", "电话", "收货地址",
        "商品名称", "数量", "供货单价", "供货小计", "客户支付",
        "真金白银(整单)", "储值抵扣(整单)", "支付方式",
        "发货状态", "退款金额", "结算状态",
    ]
    ws.append(headers)
    hfill = PatternFill("solid", fgColor="1F6FEB")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = hfill
        c.alignment = Alignment(horizontal="center", vertical="center")

    total_supply = 0.0
    total_cash = 0.0
    total_sv = 0.0
    for o in orders:
        if not o.is_refunded:
            total_cash += float(o.cash_paid or 0)
            total_sv += float(o.stored_value_paid or 0)
        items = list(o.items or [])
        if supply_only:
            items = [it for it in items if (it.supply_price is not None or it.product_id is not None)]
        if not items:
            continue
        bj = (o.order_date + timedelta(hours=8)).strftime("%Y-%m-%d %H:%M") if o.order_date else ""
        settle_label = "已结算" if o.settlement_id else "未结算"
        ship_label = _SHIP_LABEL.get(getattr(o.shipping_status, "value", o.shipping_status), "")
        cash = float(o.cash_paid or 0)
        sv = float(o.stored_value_paid or 0)
        if cash > 0 and sv > 0:
            pay_method = "混合"
        elif sv > 0:
            pay_method = "储值"
        else:
            pay_method = "现金"
        for idx, it in enumerate(items):
            sub = float(it.supply_subtotal or 0)
            if not o.is_refunded:
                total_supply += sub
            retail = (float(it.retail_price) * it.quantity) if it.retail_price else None
            # 真金白银/储值是整单金额，只在该订单第一行显示，避免多商品行重复累计
            first = idx == 0
            ws.append([
                o.wemall_order_id, bj, o.buyer_name or "", o.buyer_phone or "",
                o.shipping_address or "", it.product_name or "", it.quantity,
                float(it.supply_price) if it.supply_price else None,
                sub if it.supply_subtotal else None,
                round(retail, 2) if retail else None,
                (cash if cash else None) if first else None,
                (sv if sv else None) if first else None,
                pay_method if first else "",
                ship_label,
                float(o.refund_amount) if o.is_refunded and o.refund_amount else None,
                settle_label,
            ])

    # 合计行（col9=供货小计, col11=真金白银, col12=储值抵扣；均不含退款单）
    ws.append([])
    total_row = ["合计(不含退款)", "", "", "", "", "", "", "", round(total_supply, 2),
                 "", round(total_cash, 2), round(total_sv, 2)]
    ws.append(total_row)
    last = ws.max_row
    ws.cell(last, 1).font = Font(bold=True)
    ws.cell(last, 9).font = Font(bold=True, color="D46B08")
    ws.cell(last, 11).font = Font(bold=True, color="067647")
    ws.cell(last, 12).font = Font(bold=True, color="B54708")

    widths = [20, 17, 10, 14, 40, 36, 6, 10, 11, 11, 13, 13, 9, 10, 11, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"orders-{(start_date or datetime.now()).strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.get("/stats")
def get_order_stats(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    from sqlalchemy import func
    from app.models.order import OrderItem
    from decimal import Decimal

    q = db.query(Order).filter(_active_store_filter(db))
    q = q.filter(Order.is_test == False)  # 统计排除测试订单（含计数）
    if start_date:
        q = q.filter(Order.order_date >= start_date)
    if end_date:
        q = q.filter(Order.order_date <= end_date)

    from app.models.settlement import Settlement, SettlementStatus

    orders = q.options(joinedload(Order.items)).all()
    total_supply = sum(
        sum(item.supply_subtotal or 0 for item in o.items)
        for o in orders if not o.is_refunded and not o.is_test
    )
    total_refund = sum(o.refund_amount or 0 for o in orders if o.is_refunded and not o.is_test)

    # 未结算：没有 settlement_id 的订单
    unsettled = sum(
        sum(item.supply_subtotal or 0 for item in o.items)
        for o in orders if not o.is_refunded and not o.settlement_id and not o.is_test
    )

    # 已确认收款（真正结清）：settlement.status = 'settled'
    settlement_ids = list({o.settlement_id for o in orders if o.settlement_id})
    confirmed_ids = set()
    if settlement_ids:
        confirmed_ids = {
            s.id for s in db.query(Settlement.id)
            .filter(Settlement.id.in_(settlement_ids), Settlement.status == SettlementStatus.settled)
            .all()
        }
    confirmed_settled = sum(
        sum(item.supply_subtotal or 0 for item in o.items)
        for o in orders if not o.is_refunded and not o.is_test and o.settlement_id in confirmed_ids
    )

    return {
        "total_orders": len(orders),
        "total_supply_rmb": float(total_supply),
        "total_refund_rmb": float(total_refund),
        "net_supply_rmb": float(total_supply - total_refund),
        "unsettled_rmb": float(unsettled),
        "confirmed_settled_rmb": float(confirmed_settled),
        # 在结算单中但未确认收款
        "pending_settlement_rmb": float(total_supply - unsettled - confirmed_settled),
    }


@router.get("/today")
def today_stats(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    """今日统计（按北京时区，排除测试单）。
    order_date 存的是 naive UTC，北京今日 00:00 = UTC 今日-8h，据此过滤避免漏掉北京 0-8 点的单。
    返回今日：订单数 / 供货额 / 真金白银 / 储值抵扣 / 退款额。
    """
    from datetime import timezone
    bj = timezone(timedelta(hours=8))
    bj_now = datetime.now(bj)
    bj_midnight = bj_now.replace(hour=0, minute=0, second=0, microsecond=0)
    start_utc = (bj_midnight - timedelta(hours=8)).replace(tzinfo=None)  # naive UTC

    orders = (
        db.query(Order)
        .filter(_active_store_filter(db), Order.is_test == False, Order.order_date >= start_utc)
        .options(joinedload(Order.items))
        .all()
    )
    order_count = sum(1 for o in orders if not o.is_refunded)
    supply = sum(
        sum(item.supply_subtotal or 0 for item in o.items)
        for o in orders if not o.is_refunded
    )
    cash = sum(float(o.cash_paid or 0) for o in orders if not o.is_refunded)
    sv = sum(float(o.stored_value_paid or 0) for o in orders if not o.is_refunded)
    refund = sum(float(o.refund_amount or 0) for o in orders if o.is_refunded)
    return {
        "date": bj_now.strftime("%Y-%m-%d"),
        "order_count": order_count,
        "supply_rmb": float(supply),
        "cash_rmb": round(cash, 2),
        "stored_value_rmb": round(sv, 2),
        "refund_rmb": round(refund, 2),
    }


@router.get("/cash-daily")
def cash_daily(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """每日真金白银(在线实付)统计。按北京日期分组，排除测试单。
    cash=真金白银, stored_value=储值抵扣, full_sv_count=全储值单(现金为0)。
    退款单的现金单独计入 refund_cash，不从 cash 里扣（cash=毛到账）。
    """
    from collections import defaultdict
    q = db.query(Order).filter(_active_store_filter(db), Order.is_test == False)
    if start_date:
        q = q.filter(Order.order_date >= start_date)
    if end_date:
        q = q.filter(Order.order_date <= end_date)
    orders = q.all()

    day = defaultdict(lambda: {"cash": 0.0, "stored_value": 0.0, "order_count": 0,
                               "full_sv_count": 0, "refund_cash": 0.0})
    for o in orders:
        bj = (o.order_date + timedelta(hours=8)).strftime("%Y-%m-%d") if o.order_date else "?"
        cash = float(o.cash_paid or 0)
        sv = float(o.stored_value_paid or 0)
        d = day[bj]
        d["order_count"] += 1
        if o.is_refunded:
            d["refund_cash"] += cash
        else:
            d["cash"] += cash
            d["stored_value"] += sv
            if cash == 0 and sv > 0:
                d["full_sv_count"] += 1

    days = [{"date": k, **{kk: (round(vv, 2) if isinstance(vv, float) else vv) for kk, vv in v.items()}}
            for k, v in sorted(day.items())]
    total_cash = round(sum(d["cash"] for d in days), 2)
    total_sv = round(sum(d["stored_value"] for d in days), 2)
    total_refund = round(sum(d["refund_cash"] for d in days), 2)
    return {
        "days": days,
        "total_cash": total_cash,
        "total_stored_value": total_sv,
        "total_refund_cash": total_refund,
        "net_cash": round(total_cash - total_refund, 2),
    }


@router.post("/bulk-mark-test")
def bulk_mark_test(
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    """把当前激活店铺所有未结算订单标记为测试订单（不计入结算）"""
    store_filter = _active_store_filter(db)
    count = (
        db.query(Order)
        .filter(store_filter, Order.settlement_id == None)
        .update({"is_test": True}, synchronize_session=False)
    )
    db.commit()
    return {"updated": count}


@router.get("/{order_id}", response_model=OrderResponse)
def get_order(order_id: int, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    order = db.query(Order).options(joinedload(Order.items)).filter(Order.id == order_id, _active_store_filter(db)).first()
    if not order:
        raise HTTPException(status_code=404, detail="订单不存在")
    return order


@router.put("/{order_id}", response_model=OrderResponse)
def update_order(
    order_id: int,
    payload: OrderUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),  # 编辑订单仅管理员（运营/对账人员不可改 is_test/退款等影响结算的字段）
):
    order = db.query(Order).filter(Order.id == order_id, _active_store_filter(db)).first()
    if not order:
        raise HTTPException(status_code=404, detail="订单不存在")
    for k, v in payload.model_dump(exclude_none=True).items():
        setattr(order, k, v)
    db.commit()
    db.refresh(order)
    return order


@router.post("/sync-wemall")
async def sync_orders_from_wemall(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),  # 同步微盟订单仅管理员
):
    """从微盟拉取订单并入库"""
    from app.services.order_sync import sync_orders
    result = await sync_orders(db, start_date, end_date)
    return result
